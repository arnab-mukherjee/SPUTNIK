import pandas as pd
import numpy as np
import os
import csv
import glob
from openpyxl import load_workbook
import warnings

# Function to input week number, and identify input PDL location and output file path
def week_number():

    weekNum = input("Enter Week #: ")

    inputPDL_path = "./Input_PDL/2019-W" + weekNum + " HF Saturday PDL.csv"

    outputFile_path = "./Output/Core_Slotting_W" + weekNum + "_Prelim.xlsx"
    
    return inputPDL_path, outputFile_path

# Function to open PDL file and extract data required for slotting tool
def read_PDL(inputPDL_path):

    print("Reading PDL File...")
    input_PDL = pd.read_csv(inputPDL_path)
    
    print("Extracting PDL data...")
    colNames = ['product','bag_size','meal_swap','ice_string','box_type','inserts']
    df_ODL = input_PDL[colNames]
    
    return df_ODL

#Function to export the PDL data to output file (Slotting Tool) ODL Tab
def tool_ODL(outputFile_path, df_ODL_complete):

    print("Exporting PDL and Ingredients data to Slotting Tool (ODL Tab)...")
    
    for name in glob.glob('./Tool/Core - Slotting Tool*.xlsx'):
        tool_path = name

    book = load_workbook(tool_path, read_only=False)

    writer = pd.ExcelWriter(outputFile_path, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_ODL_complete.to_excel(writer, sheet_name='ODL', startrow=1, index=False, header=False)

    writer.save()
    writer.close()

# Function to read MARVIN's ADL output of Product Creation Tool 
def read_PCT():
    
    print("Reading Product List...")

    for name in glob.glob('./Product_List/*ADL*.csv'):
        pct_path = name

    input_PCT = pd.read_csv(pct_path)

    colNames = ['recipe_num', 'prod_id','prod_description']
    df_prodList = input_PCT[colNames]

    df_prodList = df_prodList.drop_duplicates(subset='prod_id')
    
    return df_prodList

# Function to export Product List to Slotting Tool (SKU List Tab)
def tool_SKU(outputFile_path, df_prodList):

    print("Exporting Product Creation Tool Data to Slotting Tool (SKU List Tab)...")

    book = load_workbook(outputFile_path, read_only=False)

    writer = pd.ExcelWriter(outputFile_path, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_prodList.to_excel(writer, sheet_name='SKU List', startrow=1, index=False, header=False)
    
    writer.save()
    writer.close()
    
def populate_ODL(df_ODL, df_prodList):

    print("Generating ODL Tab data from PDL...")
    
    #Convert box_type to string
    df_ODL.box_type = df_ODL.box_type.astype(str)
    
    #Populate Number of Meals column as first character of box_type string 
    df_ODL["# of meals"] = df_ODL['box_type'].str[0]
    df_ODL["# of meals"] = pd.to_numeric(df_ODL["# of meals"], errors='coerce')
    
    
    #Split meal swaps into individual components into an array
    df_ODL["meal_swap_split"] = df_ODL['meal_swap'].str.split()
    
    #Populate Meal Numbers from the split array
    df_ODL["Meal 1"] = df_ODL["meal_swap_split"].str[0]
    df_ODL["Meal 2"] = df_ODL["meal_swap_split"].str[1]
    df_ODL["Meal 3"] = df_ODL["meal_swap_split"].str[2]
    df_ODL["Meal 4"] = df_ODL["meal_swap_split"].str[3]
    
    #Extract proteins from Product List using prefix "DROP"
    df_proteins = df_prodList[df_prodList['prod_description'].str.contains('DROP')]
    
    #Convert recipe_num to int
    df_proteins["recipe_num"] = df_proteins["recipe_num"].astype(int)
    
    #Convert Meal 1, 2, 3, 4 columns to type float
    convert_columns = ["Meal 1", "Meal 2", "Meal 3", "Meal 4"]
    df_ODL[convert_columns] = df_ODL[convert_columns].astype(float)
    
    #Merge protein list with ODL to populate Protein 1
    df_ODL = pd.merge(df_ODL, df_proteins[["recipe_num","prod_description"]], left_on=['Meal 1'], right_on=['recipe_num'], how='left')
    df_ODL.drop("recipe_num", axis=1, inplace=True)
    df_ODL.rename(columns={'prod_description':'Pro 1'}, inplace=True)
    
    #Merge protein list with ODL to populate Protein 2
    df_ODL = pd.merge(df_ODL, df_proteins[["recipe_num","prod_description"]], left_on=['Meal 2'], right_on=['recipe_num'], how='left')
    df_ODL.drop("recipe_num", axis=1, inplace=True)
    df_ODL.rename(columns={'prod_description':'Pro 2'}, inplace=True)
    
    #Merge protein list with ODL to populate Protein 3
    df_ODL = pd.merge(df_ODL, df_proteins[["recipe_num","prod_description"]], left_on=['Meal 3'], right_on=['recipe_num'], how='left')
    df_ODL.drop("recipe_num", axis=1, inplace=True)
    df_ODL.rename(columns={'prod_description':'Pro 3'}, inplace=True)
    
    #Merge protein list with ODL to populate Protein 4
    df_ODL = pd.merge(df_ODL, df_proteins[["recipe_num","prod_description"]], left_on=['Meal 4'], right_on=['recipe_num'], how='left')
    df_ODL.drop("recipe_num", axis=1, inplace=True)
    df_ODL.rename(columns={'prod_description':'Pro 4'}, inplace=True)

    
    #Populate Meal Numbers from the split array
    df_ODL["Meal 1"] = df_ODL["meal_swap_split"].str[0]
    df_ODL["Meal 2"] = df_ODL["meal_swap_split"].str[1]
    df_ODL["Meal 3"] = df_ODL["meal_swap_split"].str[2]
    df_ODL["Meal 4"] = df_ODL["meal_swap_split"].str[3]
    
    #Convert Meal columns to numeric type
    df_ODL["Meal 1"] = pd.to_numeric(df_ODL["Meal 1"], errors='coerce')
    df_ODL["Meal 2"] = pd.to_numeric(df_ODL["Meal 2"], errors='coerce')
    df_ODL["Meal 3"] = pd.to_numeric(df_ODL["Meal 3"], errors='coerce')
    df_ODL["Meal 4"] = pd.to_numeric(df_ODL["Meal 4"], errors='coerce')
 
    #Split ice_string into individual elements into an array
    iceStack = pd.DataFrame(df_ODL['ice_string'].str.split("_"))
    
    #Populate Bottom and Middle string from split string   
    df_ODL["Bottom"] = iceStack["ice_string"].str[0]
    df_ODL["Middle"] = iceStack["ice_string"].str[1]
    
    #Populate Bottom Ice and Middle Ice column from first 2 characters of Bottom and Middle
    df_ODL["Bottom Ice 1"] = df_ODL["Bottom"].str[0:2]
    df_ODL["Bottom Ice 2"] = np.nan
    df_ODL["Middle Ice 1"] = df_ODL["Middle"].str[0:2]
    df_ODL["Middle Ice 2"] = np.nan
    
    #Populate CHIP column with CHIP in all rows
    df_ODL["CHIP"] = "CHIP"
    
    #Split inserts column elements into an array
    df_ODL["inserts_split"] = df_ODL['inserts'].str.split(" ")
    
    #Populate A, B, C, D column based on inserts split array
    df_ODL.inserts_split = df_ODL.inserts_split.astype(str)
    df_ODL["A"] = df_ODL["inserts_split"].apply(lambda x: 1 if "A1" in x else (2 if "A2" in x else ""))
    df_ODL["B"] = df_ODL["inserts_split"].apply(lambda x: 1 if "B1" in x else (2 if "B2" in x else ""))
    df_ODL["C"] = df_ODL["inserts_split"].apply(lambda x: 1 if "C1" in x else (2 if "C2" in x else ""))
    df_ODL["D"] = df_ODL["inserts_split"].apply(lambda x: 1 if "D1" in x else (2 if "D2" in x else ""))
    df_ODL["E"] = df_ODL["inserts_split"].apply(lambda x: 1 if "E1" in x else (2 if "E2" in x else ""))
    df_ODL["G"] = df_ODL["inserts_split"].apply(lambda x: 1 if "G1" in x else (2 if "G2" in x else ""))
    
    #Populate Top Ice column based on last 3 characters of ice string
    df_ODL["Top Ice"] = df_ODL["ice_string"].apply(lambda x: "T3F" if x[-3:] == "T3F" else ("T5F" if x[-3:] == "T5F" else ""))
    
    #Populate Recipe Cards based on Meals 
    df_ODL["Recipe Card 1"] = df_ODL["Meal 1"]
    df_ODL["Recipe Card 2"] = df_ODL["Meal 2"]
    df_ODL["Recipe Card 3"] = df_ODL["Meal 3"]
    df_ODL["Recipe Card 4"] = df_ODL["Meal 4"]
    
    #Populate HP column with HP for all rows
    df_ODL["HP"] = "HP"
    
    #Popualate Freebies and Welcome column based on inserts split array
    df_ODL["Freebies"] = df_ODL["inserts_split"].apply(lambda x: "F" if "F" in x else "")
    df_ODL["Welcome (N)"] = df_ODL["inserts_split"].apply(lambda x: "N" if "N" in x else "")
    
    #Populate Media Add-on columns based on inserts split array
    df_ODL["Media Add-on A"] = df_ODL["inserts_split"].apply(lambda x: "A" if "A1" in x else ("A" if "A2" in x else ""))
    df_ODL["Media Add-on B"] = df_ODL["inserts_split"].apply(lambda x: "B" if "B1" in x else ("B" if "B2" in x else ""))
    df_ODL["Media Add-on C"] = df_ODL["inserts_split"].apply(lambda x: "C" if "C1" in x else ("C" if "C2" in x else ""))
    df_ODL["Media Add-on D"] = df_ODL["inserts_split"].apply(lambda x: "D" if "D1" in x else ("D" if "D2" in x else ""))
    
    #Populate Brand Partnership column based on inserts split array
    df_ODL["P"] = df_ODL["inserts_split"].apply(lambda x: "P" if "P" in x else "")
    df_ODL["PI"] = df_ODL["inserts_split"].apply(lambda x: "PI" if "PI" in x else "")
    df_ODL["Z1"] = df_ODL["inserts_split"].apply(lambda x: "Z1" if "Z1" in x else "")
    
    #Drop meal_swap_split and inserts_split column from final dataframe
    df_ODL = df_ODL.drop(["meal_swap_split","inserts_split"], axis=1)
    
    return df_ODL

if __name__ == '__main__':
    
    warnings.filterwarnings("ignore")
    
    inputPDL_path, outputFile_path = week_number()
    df_ODL = read_PDL(inputPDL_path)
    df_prodList = read_PCT()
    df_ODL_complete = populate_ODL(df_ODL, df_prodList)
    tool_ODL(outputFile_path, df_ODL_complete)
    tool_SKU(outputFile_path, df_prodList)
    print("Finish.")

